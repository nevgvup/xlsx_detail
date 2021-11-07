# -*- coding: utf-8 -*-
import openpyxl

class handle_excel:
    sheet_cur_name = "null"         # 当前活动sheet页
    xlsx_file_path = "null"         # 文件路径
    def __init__(self, file_path, index = None):
        '''
        加载excel，入参文件名、当前活动的sheet页
        返回当前文件
        '''
        self.xlsx_wb = openpyxl.load_workbook(file_path) # 拿到excel的所有内容
        self.sheet_name = self.xlsx_wb.sheetnames # 拿到sheetnames的所有内容
        self.set_cur_sheet(index)
        print("file name:", file_path, "cur sheet:", self.sheet_cur_name)
        self.xlsx_file_path = file_path
    def set_cur_sheet(self, index = None):
        '''
        配置当前活动的sheet页
        返回当前活动的sheet页
        '''
        if index == None:
            index = 0
        self.sheet_cur_name = self.sheet_name[index]
        self.cur_sheet = self.xlsx_wb[self.sheet_cur_name]

    def get_cell_value(self, row, cols):
        '''
        获取某一个单元格内容
        '''
        data = self.cur_sheet.cell(row = row, column = cols)
        return data
    def set_cell_value(self, row, cols, value):
        '''
        设置某一个单元格内容
        '''
        data = self.cur_sheet.cell(row = row, column = cols)
        data.value = value
    def get_col_max(self):
        '''
        获取当前活动页的最大列数
        '''
        col_num = 0
        for col in self.cur_sheet.iter_cols():
            col_num += 1
        print("max_column:", self.cur_sheet.max_column, "col_num:", col_num)
        return col_num
    def get_row_list(self,row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.cur_sheet[row]:
            row_list.append(i.value)
        return row_list
    def get_col_num_from_row(self, row, value):
        '''
        获取获取指定行(行号从1开始) 指定数据数据的列号
        成功返回对应的列号，失败返回None
        '''
        col_num = None
        num = 1
        for cell in self.cur_sheet[row]:
            if cell.value == value:
                col_num = num
                break
            num += 1
        return col_num
    def copy_row(self, row, num = 1):
        '''
        复制多行
        '''
        for i in range(num):
            col_num = 0
            new_row = row + i + 1
            self.cur_sheet.insert_rows(new_row)
            for cell in self.cur_sheet[row]:
                col_num += 1
                # print("col_num", col_num, "cell.value:", cell.value)
                self.set_cell_value(new_row, col_num, cell.value)

        self.xlsx_wb.save(self.xlsx_file_path)

    # 行操作
    def get_row_max(self):
        '''
        获取当前活动页的最大行数
        '''
        row_num = 0
        for row in self.cur_sheet.iter_rows():
            row_num += 1
        return row_num
    def insert_col(self, col_num, title):
        '''
        插入一列,附带标题
        '''
        self.cur_sheet.insert_cols(col_num)
        self.set_cell_value(1, col_num, title)
        self.xlsx_wb.save(self.xlsx_file_path)
    def save(self):
        self.xlsx_wb.save(self.xlsx_file_path)